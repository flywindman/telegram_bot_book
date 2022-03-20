from openpyxl import Workbook
import datetime
"""
# создание эксель файла
#__________________________________________________________________________________
excel_file = Workbook()
excel_sheet = excel_file.create_sheet(title='Holidays 2019', index=0)

# создадим заголовки столбцов
excel_sheet['A1'] = 'Holiday Name'
excel_sheet['B1'] = 'Holiday Description'
excel_sheet['C1'] = 'Holiday Date'

# добавим данные
excel_sheet['A2'] = 'Diwali'
excel_sheet['B2'] = 'Biggest Indian Festival'
excel_sheet['C2'] = datetime.date(year=2019, month=10, day=27).strftime("%m/%d/%y")

excel_sheet['A3'] = 'Christmas'
excel_sheet['B3'] = 'Birth of Jesus Christ'
excel_sheet['C3'] = datetime.date(year=2019, month=12, day=25).strftime("%m/%d/%y")

# сохраним файл
excel_file.save(filename="Holidays.xlsx")
#__________________________________________________________________________________________
"""

"""
Мы можем либо использовать индекс ячейки или использовать объект ячейки для установки значения. 
Изменим некоторые значения в файле Excel, созданные в последнем примере
"""
"""
import openpyxl

excel_file = openpyxl.load_workbook('Holidays.xlsx')
excel_sheet = excel_file['Holidays 2019']

# перезаписываем ячейку используя ее индекс
excel_sheet['A2'] = 'Deepawali'

# и используя объект ячейка
excel_sheet.cell(row=2, column=2).value = 'Biggest Indian Festival for Hindus'

excel_file.save('Holidays.xlsx')
#__________________________________________________________________________________________
"""

"""
Добавление нескольких значений в лист Excel
"""
import openpyxl
excel_file = openpyxl.load_workbook('Holidays.xlsx')
excel_sheet = excel_file['Holidays 2019']

holiday_rows = (
    ('Black Friday', 'Fourth Thursday of November, Shopping Day', '11/29/19'),
    ('Holi', 'Festival of Colors', '3/20/19')
)

for row in holiday_rows:
    excel_sheet.append(row)

excel_file.save('Holidays.xlsx')
